import os
import json
import excel_backup
import pdf_challan
from flask import send_from_directory
from flask import Flask, render_template, request, redirect, flash, session
from db import get_db_connection, log_change, log_notification
from datetime import datetime, timedelta
from rapidfuzz import fuzz, process


app = Flask(__name__)
app.secret_key = "krupalu-dashboard-secret-key"

@app.route("/")
def dashboard():
    conn = get_db_connection()

    period = request.args.get("period", "30")

    date_filter_sql = ""
    if period != "all":
        cutoff = (datetime.now() - timedelta(days=int(period))).strftime("%Y-%m-%d")
        date_filter_sql = f"AND ic.date_received >= '{cutoff}'"

    order_rows = conn.execute(f"""
        SELECT p.company_name AS company, p.phone AS phone, ii.fabric_name AS fabric,
               ii.fabric_width_inches AS width, gt.name AS garment_type,
               ii.meters_given AS given,
               COALESCE(SUM(d.meters_delivered), 0) AS printing_total,
               COALESCE((
                   SELECT SUM(dt.meters) FROM delivery_takkas dt
                   JOIN deliveries d2 ON dt.delivery_id = d2.id
                   WHERE d2.inward_item_id = ii.id
               ), 0) AS takka_total
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        LEFT JOIN garment_types gt ON gt.id = ii.garment_type_id
        LEFT JOIN deliveries d ON d.inward_item_id = ii.id
        WHERE 1=1 {date_filter_sql}
        GROUP BY ii.id
        ORDER BY ii.id DESC
    """).fetchall()

    orders = []
    for o in order_rows:
        orders.append({
            "company": o["company"],
            "phone": o["phone"],
            "fabric": o["fabric"],
            "size": f'{o["width"]}"' if o["width"] else "-",
            "garment_type": o["garment_type"] or "-",
            "given": o["given"],
            "delivered": round(o["takka_total"], 2),
            "shortage": round(o["printing_total"] - o["takka_total"], 2),
            "balance": round(o["given"] - o["printing_total"], 2)
        })

    orders_display = [o for o in orders if o["balance"] > 0]
    total_meters = sum(o["given"] for o in orders)
    total_delivered = sum(o["given"] - o["balance"] for o in orders)

    shortage_meters = round(sum(o["shortage"] for o in orders), 2)
    avg_shortage = round((shortage_meters / total_meters * 100), 1) if total_meters > 0 else 0

    pending_total = conn.execute(f"""
        SELECT COALESCE(SUM(pay.total_amount - pay.amount_paid), 0) AS total
        FROM payments pay
        JOIN deliveries d ON pay.delivery_id = d.id
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        WHERE 1=1 {date_filter_sql}
    """).fetchone()["total"]

    conn.close()

    return render_template(
        "dashboard.html",
        total_meters=total_meters,
        total_delivered=total_delivered,
        payment_pending=round(pending_total, 2),
        avg_shortage=avg_shortage,
        shortage_meters=shortage_meters,
        orders=orders_display,
        selected_period=period
    )

@app.route("/new-work", methods=["GET", "POST"])
def new_work():
    conn = get_db_connection()

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        phone = request.form.get("phone", "").strip()
        challan_number = request.form.get("challan_number", "").strip()
        date_received = request.form.get("date_received", "").strip()

        if not (company_name and phone and challan_number and date_received):
            conn.close()
            return "Company, phone, challan number, aur date bharna zaroori hai.", 400

        if not (phone.isdigit() and len(phone) == 10):
            conn.close()
            return "Phone number sirf 10 digit ka hona chahiye.", 400

        fabric_names = request.form.getlist("fabric_name")
        widths = request.form.getlist("fabric_width_inches")
        garment_ids = request.form.getlist("garment_type_id")
        takkas = request.form.getlist("total_takka")
        meters_list = request.form.getlist("meters_given")
        rates = request.form.getlist("rate_per_meter")

        garment_rows = conn.execute("SELECT id, name FROM garment_types").fetchall()
        garment_name_map = {g["id"]: g["name"] for g in garment_rows}

        party = conn.execute(
            "SELECT id FROM parties WHERE company_name = ? AND phone = ?",
            (company_name, phone)
        ).fetchone()

        if party:
            party_id = party["id"]
        else:
            cur = conn.execute(
                "INSERT INTO parties (company_name, phone) VALUES (?, ?)",
                (company_name, phone)
            )
            party_id = cur.lastrowid

        cur = conn.execute(
            "INSERT INTO inward_challans (party_id, inward_challan_number, date_received) VALUES (?, ?, ?)",
            (party_id, challan_number, date_received)
        )
        challan_id = cur.lastrowid

        saved_count = 0
        for i in range(len(fabric_names)):
            fabric_name = fabric_names[i].strip() if i < len(fabric_names) else ""
            meters_given = meters_list[i].strip() if i < len(meters_list) else ""
            if not fabric_name or not meters_given:
                continue

            width = widths[i] if i < len(widths) and widths[i] else None
            garment_id = garment_ids[i] if i < len(garment_ids) and garment_ids[i] else None
            takka = takkas[i] if i < len(takkas) and takkas[i] else None
            rate = rates[i] if i < len(rates) and rates[i] else None

            conn.execute(
                """INSERT INTO inward_items
                   (inward_challan_id, fabric_name, garment_type_id, meters_given,
                    rate_per_meter, fabric_width_inches, total_takka)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (challan_id, fabric_name, garment_id, meters_given, rate, width, takka)
            )
            conn.execute("INSERT OR IGNORE INTO fabric_master (fabric_name) VALUES (?)", (fabric_name,))

            garment_name = garment_name_map.get(int(garment_id)) if garment_id else "-"

            excel_backup.append_order(
                company_name, phone, date_received, challan_number,
                fabric_name, width, garment_name, takka, meters_given, rate
            )
            saved_count += 1

        log_notification(conn, f"Order submitted for {company_name} ({saved_count} fabric(s))")
        conn.commit()
        conn.close()
        flash(f"Order for {company_name} added successfully - {saved_count} fabric(s)")
        return redirect("/")

    garment_types = conn.execute("SELECT * FROM garment_types").fetchall()
    conn.close()

    with open("config.json") as f:
        config = json.load(f)
    fabric_widths = config.get("fabric_widths_inches", [])

    return render_template("new_work.html", garment_types=garment_types, fabric_widths=fabric_widths)

@app.route("/deliveries", methods=["GET", "POST"])
def deliveries():
    conn = get_db_connection()

    if request.method == "POST":
        outward_challan_number = request.form.get("outward_challan_number", "").strip()
        delivery_date = request.form.get("delivery_date", "").strip()

        item_ids = request.form.getlist("item_id")
        printing_list = request.form.getlist("printing_meters")

        if not (outward_challan_number and delivery_date and item_ids):
            conn.close()
            return "Challan number, date, and at least one fabric entry are required.", 400

        saved_count = 0
        for i in range(len(item_ids)):
            takka_raw = request.form.get(f"takka_list_{item_ids[i]}", "").strip()
            if not takka_raw:
                continue

            takka_values = [float(t) for t in takka_raw.split(",") if t.strip()]
            takka_total = sum(takka_values)
            if takka_total <= 0:
                continue

            printing_meters = printing_list[i].strip() if i < len(printing_list) else ""
            printing_meters = float(printing_meters) if printing_meters else takka_total

            # For balance tracking, printing_meters acts as the total meters delivered
            cur = conn.execute(
                """INSERT INTO deliveries
                   (inward_item_id, outward_challan_number, meters_delivered, delivery_date)
                   VALUES (?, ?, ?, ?)""",
                (item_ids[i], outward_challan_number, printing_meters, delivery_date)
            )
            delivery_id = cur.lastrowid

            for t in takka_values:
                conn.execute(
                    "INSERT INTO delivery_takkas (delivery_id, meters) VALUES (?, ?)",
                    (delivery_id, t)
                )

            # Shortage is handled completely automatically by the system backend
            shortage = round(printing_meters - takka_total, 2)
            if shortage > 0:
                conn.execute(
                    """INSERT INTO fabric_damage
                       (inward_item_id, damage_meters, stage, reported_by, note)
                       VALUES (?, ?, 'delivery', 'system', 'Auto: printing meters minus takka total')""",
                    (item_ids[i], shortage)
                )
            saved_count += 1

        conn.commit()

        for i in range(len(item_ids)):
            takka_raw = request.form.get(f"takka_list_{item_ids[i]}", "").strip()
            if not takka_raw:
                continue
            takka_values = [float(t) for t in takka_raw.split(",") if t.strip()]
            takka_total = sum(takka_values)
            if takka_total <= 0:
                continue

            printing_meters = printing_list[i].strip() if i < len(printing_list) else ""
            printing_meters = float(printing_meters) if printing_meters else takka_total
            shortage = round(printing_meters - takka_total, 2)

            item_info = conn.execute("""
                SELECT p.company_name, p.phone, ii.fabric_name, ii.fabric_width_inches
                FROM inward_items ii
                JOIN inward_challans ic ON ii.inward_challan_id = ic.id
                JOIN parties p ON ic.party_id = p.id
                WHERE ii.id = ?
            """, (item_ids[i],)).fetchone()

            takka_count = len(takka_values)

            excel_backup.append_delivery(
                item_info["company_name"], item_info["phone"], delivery_date, outward_challan_number,
                item_info["fabric_name"], item_info["fabric_width_inches"], takka_total, takka_count, printing_meters, shortage
            )

        conn.close()

        conn2 = get_db_connection()
        log_notification(conn2, f"Delivery challan {outward_challan_number} created")
        conn2.commit()
        conn2.close()

        party_id = request.form.get("party_id_for_whatsapp")
        return redirect(f"/deliveries/confirm/{outward_challan_number}?party_id={party_id}")

    # FIXED: Only fetch companies that match the active dashboard entries (balance > 0)
    parties = conn.execute("""
        SELECT DISTINCT p.id, p.company_name, p.phone 
        FROM parties p
        JOIN inward_challans ic ON ic.party_id = p.id
        JOIN inward_items ii ON ii.inward_challan_id = ic.id
        LEFT JOIN deliveries d ON d.inward_item_id = ii.id
        GROUP BY ii.id
        HAVING (ii.meters_given - COALESCE(SUM(d.meters_delivered), 0)) > 0
        ORDER BY p.company_name
    """).fetchall()

    rows = conn.execute("""
        SELECT ii.id, p.company_name, ii.fabric_name, ii.fabric_width_inches, ii.meters_given,
               COALESCE(SUM(d.meters_delivered), 0) AS delivered
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        LEFT JOIN deliveries d ON d.inward_item_id = ii.id
        GROUP BY ii.id
        ORDER BY ii.id DESC
    """).fetchall()

    all_items = []
    for r in rows:
        balance = r["meters_given"] - r["delivered"]
        if balance <= 0:
            continue
        all_items.append({
            "company_name": r["company_name"],
            "fabric_name": r["fabric_name"],
            "fabric_width_inches": r["fabric_width_inches"] or "-",
            "meters_given": r["meters_given"],
            "delivered": r["delivered"],
            "balance": balance
        })

    conn.close()
    return render_template("deliveries.html", parties=parties, all_items=all_items)

@app.route("/payments", methods=["GET", "POST"])
def payments():
    conn = get_db_connection()

    if request.method == "POST":
        delivery_id = request.form["delivery_id"]
        base_amount = float(request.form["base_amount"])
        gst_percent = float(request.form["gst_percent"])
        due_days = int(request.form["due_days"])

        gst_amount = base_amount * gst_percent / 100
        total_amount = base_amount + gst_amount

        delivery = conn.execute(
            "SELECT delivery_date FROM deliveries WHERE id = ?", (delivery_id,)
        ).fetchone()
        delivery_date = datetime.strptime(delivery["delivery_date"], "%Y-%m-%d")
        due_date = (delivery_date + timedelta(days=due_days)).strftime("%Y-%m-%d")

        conn.execute(
            """INSERT INTO payments
               (delivery_id, base_amount, gst_percent, gst_amount, total_amount, due_date, status)
               VALUES (?, ?, ?, ?, ?, ?, 'pending')""",
            (delivery_id, base_amount, gst_percent, gst_amount, total_amount, due_date)
        )
        conn.commit()
        conn.close()
        return redirect("/payments")

    # FIXED: Only fetch open deliveries that have actual quantities delivered (meters_delivered > 0)
    open_delivery_rows = conn.execute("""
        SELECT d.id, p.company_name, ii.fabric_name, d.delivery_date,
               COALESCE(SUM(dt.meters), 0) AS takka_total
        FROM deliveries d
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        LEFT JOIN delivery_takkas dt ON dt.delivery_id = d.id
        WHERE d.id NOT IN (SELECT delivery_id FROM payments)
          AND d.meters_delivered > 0
        GROUP BY d.id
        ORDER BY d.id DESC
    """).fetchall()

    open_deliveries = []
    for d in open_delivery_rows:
        y, m, day = d["delivery_date"].split("-")
        open_deliveries.append({
            "id": d["id"],
            "company_name": d["company_name"],
            "fabric_name": d["fabric_name"],
            "takka_total": round(d["takka_total"], 2),
            "delivery_date_display": f"{day}-{m}-{y}"
        })

    payment_rows = conn.execute("""
        SELECT pay.id, p.company_name, p.phone, pay.base_amount, pay.gst_percent,
               pay.total_amount, pay.amount_paid, pay.due_date, pay.status,
               d.outward_challan_number
        FROM payments pay
        JOIN deliveries d ON pay.delivery_id = d.id
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        WHERE pay.status != 'paid'
        ORDER BY pay.id DESC
    """).fetchall()

    all_payments = []
    for r in payment_rows:
        row_dict = dict(r)
        # Convert YYYY-MM-DD to DD-MM-YYYY for display
        y, m, d = row_dict["due_date"].split("-")
        row_dict["due_date_display"] = f"{d}-{m}-{y}"
        all_payments.append(row_dict)

    conn.close()
    return render_template("payments.html", open_deliveries=open_deliveries, all_payments=all_payments)

@app.route("/payments/pay/<int:payment_id>", methods=["POST"])
def add_partial_payment(payment_id):
    conn = get_db_connection()
    paid_amount = float(request.form["paid_amount"])

    conn.execute(
        "INSERT INTO payment_history (payment_id, paid_amount) VALUES (?, ?)",
        (payment_id, paid_amount)
    )

    payment = conn.execute("""
        SELECT pay.amount_paid, pay.total_amount, pay.gst_percent, p.company_name
        FROM payments pay
        JOIN deliveries d ON pay.delivery_id = d.id
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        WHERE pay.id = ?
    """, (payment_id,)).fetchone()

    new_paid = payment["amount_paid"] + paid_amount

    if new_paid >= payment["total_amount"]:
        status = "paid"
    elif new_paid > 0:
        status = "partial"
    else:
        status = "pending"

    conn.execute(
        "UPDATE payments SET amount_paid = ?, status = ? WHERE id = ?",
        (new_paid, status, payment_id)
    )
    log_notification(conn, f"Payment received from {payment['company_name']} - Rs {paid_amount}")
    conn.commit()
    conn.close()

    excel_backup.append_payment(
        payment["company_name"], datetime.now().strftime("%Y-%m-%d"),
        "Partial payment", paid_amount, payment["gst_percent"], payment["total_amount"]
    )

    return redirect("/payments")

@app.route("/parties")
def parties():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT p.id, p.company_name, p.phone,
               COALESCE(SUM(ii.meters_given), 0) AS total_given,
               COALESCE(SUM(d.meters_delivered), 0) AS total_delivered
        FROM parties p
        LEFT JOIN inward_challans ic ON ic.party_id = p.id
        LEFT JOIN inward_items ii ON ii.inward_challan_id = ic.id
        LEFT JOIN deliveries d ON d.inward_item_id = ii.id
        GROUP BY p.id
        ORDER BY p.company_name
    """).fetchall()
    conn.close()
    return render_template("parties.html", parties=rows)


@app.route("/parties/<int:party_id>")
def party_detail(party_id):
    conn = get_db_connection()
    party = conn.execute("SELECT * FROM parties WHERE id = ?", (party_id,)).fetchone()

    selected_range = request.args.get("range", "90")

    date_filter_sql = ""
    if selected_range != "all":
        days = int(selected_range)
        cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        date_filter_sql = f"AND ic.date_received >= '{cutoff}'"

    order_rows = conn.execute(f"""
        SELECT ii.id AS item_id, ii.fabric_name, ii.fabric_width_inches AS size,
               ii.rate_per_meter AS rate, ic.date_received, ic.inward_challan_number,
               ii.meters_given,
               COALESCE(pr.printing_total, 0) AS printing_total,
               COALESCE(tk.takka_total, 0) AS delivered,
               dch.outward_challans, dch.last_delivery_date,
               COALESCE(pay.total_amount, 0) AS total_amount,
               COALESCE(pay.amount_paid, 0) AS amount_paid,
               pay.status AS payment_status
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        LEFT JOIN (
            SELECT inward_item_id, SUM(meters_delivered) AS printing_total
            FROM deliveries GROUP BY inward_item_id
        ) pr ON pr.inward_item_id = ii.id
        LEFT JOIN (
            SELECT d.inward_item_id, SUM(dt.meters) AS takka_total
            FROM deliveries d
            JOIN delivery_takkas dt ON dt.delivery_id = d.id
            GROUP BY d.inward_item_id
        ) tk ON tk.inward_item_id = ii.id
        LEFT JOIN (
            SELECT inward_item_id,
                   GROUP_CONCAT(DISTINCT outward_challan_number) AS outward_challans,
                   MAX(delivery_date) AS last_delivery_date
            FROM deliveries GROUP BY inward_item_id
        ) dch ON dch.inward_item_id = ii.id
        LEFT JOIN (
            SELECT d.inward_item_id, SUM(p.total_amount) AS total_amount,
                   SUM(p.amount_paid) AS amount_paid,
                   CASE WHEN SUM(p.amount_paid) >= SUM(p.total_amount) AND SUM(p.total_amount) > 0 THEN 'paid'
                        WHEN SUM(p.amount_paid) > 0 THEN 'partial'
                        ELSE 'pending' END AS status
            FROM payments p
            JOIN deliveries d ON p.delivery_id = d.id
            GROUP BY d.inward_item_id
        ) pay ON pay.inward_item_id = ii.id
        WHERE ic.party_id = ? {date_filter_sql}
        ORDER BY ic.date_received DESC
    """, (party_id,)).fetchall()

    orders = []
    item_ids_for_history = [o["item_id"] for o in order_rows]
    payment_history_map = {}
    delivery_details_map = {}
    if item_ids_for_history:
        placeholders_d = ",".join("?" * len(item_ids_for_history))
        delivery_rows = conn.execute(f"""
            SELECT d.inward_item_id, d.outward_challan_number, d.delivery_date,
                   COALESCE((SELECT SUM(dt.meters) FROM delivery_takkas dt WHERE dt.delivery_id = d.id), 0) AS meters
            FROM deliveries d
            WHERE d.inward_item_id IN ({placeholders_d})
            ORDER BY d.delivery_date
        """, item_ids_for_history).fetchall()
        for d in delivery_rows:
            key = d["inward_item_id"]
            if key not in delivery_details_map:
                delivery_details_map[key] = []
            y, m, day = d["delivery_date"].split("-")
            delivery_details_map[key].append({
                "challan": d["outward_challan_number"],
                "date": f"{day}-{m}-{y}",
                "meters": round(d["meters"], 2)
            })
    if item_ids_for_history:
        placeholders = ",".join("?" * len(item_ids_for_history))
        history_rows = conn.execute(f"""
            SELECT d.inward_item_id, ph.paid_amount, ph.paid_date
            FROM payment_history ph
            JOIN payments p ON ph.payment_id = p.id
            JOIN deliveries d ON p.delivery_id = d.id
            WHERE d.inward_item_id IN ({placeholders})
            ORDER BY ph.paid_date
        """, item_ids_for_history).fetchall()
        for h in history_rows:
            key = h["inward_item_id"]
            if key not in payment_history_map:
                payment_history_map[key] = []
            date_part = h["paid_date"].split(" ")[0]
            y, m, d = date_part.split("-")
            payment_history_map[key].append({"date": f"{d}-{m}-{y}", "amount": h["paid_amount"]})

    for o in order_rows:
        shortage = round(o["printing_total"] - o["delivered"], 2)
        remaining = round(o["total_amount"] - o["amount_paid"], 2)
        orders.append({
            "item_id": o["item_id"],
            "fabric_name": o["fabric_name"],
            "size": f'{o["size"]}"' if o["size"] else "-",
            "rate": o["rate"] if o["rate"] else "-",
            "date_received": o["date_received"],
            "inward_challan_number": o["inward_challan_number"],
            "delivery_details": delivery_details_map.get(o["item_id"], []),
            "meters_given": o["meters_given"],
            "delivered": round(o["delivered"], 2),
            "shortage": shortage,
            "total_amount": round(o["total_amount"], 2),
            "amount_paid": round(o["amount_paid"], 2),
            "remaining": remaining,
            "payment_status": o["payment_status"] or "no payment yet",
            "payment_history": payment_history_map.get(o["item_id"], [])
        })


    total_given = sum(o["meters_given"] for o in orders)
    total_delivered = sum(o["delivered"] for o in orders)
    fabric_types = ", ".join(sorted(set(o["fabric_name"] for o in orders))) or "-"

    fabric_breakdown = {}
    for o in orders:
        name = o["fabric_name"]
        fabric_breakdown[name] = fabric_breakdown.get(name, 0) + o["meters_given"]

    fabric_breakdown_list = sorted(fabric_breakdown.items(), key=lambda x: x[1], reverse=True)

    party_shortage = round(sum(o["shortage"] for o in orders), 2)
    party_shortage_pct = round((party_shortage / total_given * 100), 1) if total_given > 0 else 0

    conn.close()
    return render_template(
        "party_detail.html",
        party=party,
        orders=orders,
        selected_range=selected_range,
        total_given=total_given,
        total_delivered=total_delivered,
        fabric_types=fabric_types,
        party_shortage=party_shortage,
        party_shortage_pct=party_shortage_pct,
        fabric_breakdown_list=fabric_breakdown_list
    )

@app.route("/parties/<int:party_id>/export-excel")
def export_party_excel(party_id):
    from openpyxl import Workbook

    conn = get_db_connection()
    party = conn.execute("SELECT * FROM parties WHERE id = ?", (party_id,)).fetchone()

    rows = conn.execute("""
        SELECT ii.id AS item_id, ii.fabric_name, ii.fabric_width_inches AS size,
               ii.rate_per_meter AS rate, ic.date_received, ic.inward_challan_number,
               ii.meters_given,
               COALESCE(tk.takka_total, 0) AS delivered,
               COALESCE(pr.printing_total, 0) AS printing_total,
               COALESCE(pay.total_amount, 0) AS total_amount,
               COALESCE(pay.amount_paid, 0) AS amount_paid,
               pay.status AS payment_status
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        LEFT JOIN (
            SELECT inward_item_id, SUM(meters_delivered) AS printing_total
            FROM deliveries GROUP BY inward_item_id
        ) pr ON pr.inward_item_id = ii.id
        LEFT JOIN (
            SELECT d.inward_item_id, SUM(dt.meters) AS takka_total
            FROM deliveries d JOIN delivery_takkas dt ON dt.delivery_id = d.id
            GROUP BY d.inward_item_id
        ) tk ON tk.inward_item_id = ii.id
        LEFT JOIN (
            SELECT d.inward_item_id, SUM(p.total_amount) AS total_amount, SUM(p.amount_paid) AS amount_paid,
                   CASE WHEN SUM(p.amount_paid) >= SUM(p.total_amount) AND SUM(p.total_amount) > 0 THEN 'paid'
                        WHEN SUM(p.amount_paid) > 0 THEN 'partial' ELSE 'pending' END AS status
            FROM payments p JOIN deliveries d ON p.delivery_id = d.id
            GROUP BY d.inward_item_id
        ) pay ON pay.inward_item_id = ii.id
        WHERE ic.party_id = ?
        ORDER BY ic.date_received DESC
    """, (party_id,)).fetchall()

    item_ids = [r["item_id"] for r in rows]
    delivery_map = {}
    if item_ids:
        placeholders = ",".join("?" * len(item_ids))
        d_rows = conn.execute(f"""
            SELECT inward_item_id, outward_challan_number, delivery_date
            FROM deliveries WHERE inward_item_id IN ({placeholders})
            ORDER BY delivery_date
        """, item_ids).fetchall()
        for d in d_rows:
            delivery_map.setdefault(d["inward_item_id"], []).append(
                f"{d['outward_challan_number']} ({d['delivery_date']})"
            )

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["Company", "Fabric", "Size", "Date Received", "Inward Challan", "Outward Challan(s)",
               "Given (m)", "Delivered (m)", "Shortage (m)", "Rate/m", "Total (incl GST)",
               "Paid", "Remaining", "Status", "Delivery dates"]
    ws.append(headers)

    row_data = []
    for r in rows:
        shortage = round(r["printing_total"] - r["delivered"], 2)
        remaining = round(r["total_amount"] - r["amount_paid"], 2)
        outward = ", ".join(delivery_map.get(r["item_id"], [])) or "-"
        row_data.append([
            party["company_name"], r["fabric_name"], f'{r["size"]}"' if r["size"] else "-",
            r["date_received"], r["inward_challan_number"], outward, r["meters_given"],
            round(r["delivered"], 2), shortage, r["rate"] or "-", round(r["total_amount"], 2),
            round(r["amount_paid"], 2), remaining, r["payment_status"] or "no payment yet", outward
        ])

    for rd in row_data:
        ws.append(rd)
        ws.insert_rows(2)
        for i, val in enumerate(rd, start=1):
            ws.cell(row=2, column=i, value=val)
        ws.delete_rows(ws.max_row)

    export_folder = "excel_backups/Party_Exports"
    os.makedirs(export_folder, exist_ok=True)
    filename = pdf_challan.safe_folder_name(party["company_name"]) + "_ledger.xlsx"
    filepath = os.path.join(export_folder, filename)
    wb.save(filepath)

    return send_from_directory(os.path.abspath(export_folder), filename, as_attachment=True)

@app.route("/parties/<int:party_id>/edit-phone", methods=["POST"])
def edit_phone(party_id):
    conn = get_db_connection()
    new_phone = request.form["new_phone"]

    party = conn.execute("SELECT company_name, phone FROM parties WHERE id = ?", (party_id,)).fetchone()

    if party["phone"] != new_phone:
        summary = f"{party['company_name']} - mobile changed {party['phone']} to {new_phone}"
        log_change(conn, "party", party_id, "phone", party["phone"], new_phone, summary)
        conn.execute("UPDATE parties SET phone = ? WHERE id = ?", (new_phone, party_id))
        conn.commit()

    conn.close()
    return redirect(f"/parties/{party_id}")


@app.route("/payments/<int:payment_id>/edit-gst", methods=["POST"])
def edit_gst(payment_id):
    conn = get_db_connection()
    new_gst_percent = float(request.form["new_gst_percent"])

    payment = conn.execute("""
        SELECT pay.gst_percent, pay.base_amount, p.company_name
        FROM payments pay
        JOIN deliveries d ON pay.delivery_id = d.id
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        WHERE pay.id = ?
    """, (payment_id,)).fetchone()

    if payment["gst_percent"] != new_gst_percent:
        summary = f"{payment['company_name']} - GST changed {payment['gst_percent']}% to {new_gst_percent}%"
        log_change(conn, "payment", payment_id, "gst_percent", payment["gst_percent"], new_gst_percent, summary)

        new_gst_amount = payment["base_amount"] * new_gst_percent / 100
        new_total = payment["base_amount"] + new_gst_amount

        conn.execute(
            "UPDATE payments SET gst_percent = ?, gst_amount = ?, total_amount = ? WHERE id = ?",
            (new_gst_percent, new_gst_amount, new_total, payment_id)
        )
        conn.commit()

    conn.close()
    return redirect("/payments")


@app.route("/change-log")
def change_log():
    conn = get_db_connection()
    entries = conn.execute("SELECT * FROM audit_log ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("change_log.html", entries=entries)

@app.route("/reports")
def reports():
    return render_template("reports.html")


@app.route("/api/reports-filters")
def reports_filters():
    conn = get_db_connection()
    parties = [r["company_name"] for r in conn.execute(
        "SELECT DISTINCT company_name FROM parties ORDER BY company_name"
    ).fetchall()]
    fabrics = [r["fabric_name"] for r in conn.execute(
        "SELECT DISTINCT fabric_name FROM inward_items ORDER BY fabric_name"
    ).fetchall()]
    widths = [r["fabric_width_inches"] for r in conn.execute(
        "SELECT DISTINCT fabric_width_inches FROM inward_items WHERE fabric_width_inches IS NOT NULL ORDER BY fabric_width_inches"
    ).fetchall()]
    conn.close()
    return {"parties": parties, "fabrics": fabrics, "widths": widths}


@app.route("/api/reports-data")
def reports_data():
    from collections import defaultdict

    conn = get_db_connection()

    party = request.args.get("party", "all")
    fabric = request.args.get("fabric", "all")
    width = request.args.get("width", "all")
    range_val = request.args.get("range", "12")

    filters = []
    params = []

    if party != "all":
        filters.append("p.company_name = ?")
        params.append(party)
    if fabric != "all":
        filters.append("ii.fabric_name = ?")
        params.append(fabric)
    if width != "all":
        filters.append("ii.fabric_width_inches = ?")
        params.append(width)
    if range_val != "all":
        cutoff = (datetime.now() - timedelta(days=int(range_val) * 30)).strftime("%Y-%m-%d")
        filters.append("ic.date_received >= ?")
        params.append(cutoff)

    where_sql = ("WHERE " + " AND ".join(filters)) if filters else ""

    item_rows = conn.execute(f"""
        SELECT ii.id, ii.meters_given, ii.fabric_name, ic.date_received, p.company_name
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        {where_sql}
    """, params).fetchall()

    total_given = sum(r["meters_given"] for r in item_rows)
    item_ids = [r["id"] for r in item_rows]

    empty_response = {
        "kpis": {"given": round(total_given, 2), "delivered": 0, "shortage": 0, "pending": 0},
        "months": [], "given_by_month": [], "delivered_by_month": [],
        "fabric_labels": [], "fabric_values": [],
        "status_labels": ["Paid", "Partial", "Pending"], "status_values": [0, 0, 0],
        "party_labels": [], "party_values": []
    }

    if not item_ids:
        conn.close()
        return empty_response

    placeholders = ",".join("?" * len(item_ids))
    delivered_rows = conn.execute(f"""
        SELECT d.id, d.meters_delivered, d.delivery_date,
               COALESCE((SELECT SUM(dt.meters) FROM delivery_takkas dt WHERE dt.delivery_id = d.id), 0) AS takka_total
        FROM deliveries d
        WHERE d.inward_item_id IN ({placeholders})
    """, item_ids).fetchall()

    total_delivered = sum(r["takka_total"] for r in delivered_rows)
    total_printing = sum(r["meters_delivered"] for r in delivered_rows)
    shortage = round(total_printing - total_delivered, 2)

    delivery_ids = [r["id"] for r in delivered_rows]
    pending = 0
    status_counts = {"paid": 0, "partial": 0, "pending": 0}

    if delivery_ids:
        d_placeholders = ",".join("?" * len(delivery_ids))
        pay_rows = conn.execute(f"""
            SELECT total_amount, amount_paid, status FROM payments
            WHERE delivery_id IN ({d_placeholders})
        """, delivery_ids).fetchall()
        for pr in pay_rows:
            pending += (pr["total_amount"] - pr["amount_paid"])
            status_counts[pr["status"]] = status_counts.get(pr["status"], 0) + 1

    given_by_month = defaultdict(float)
    for r in item_rows:
        given_by_month[r["date_received"][:7]] += r["meters_given"]

    delivered_by_month = defaultdict(float)
    for r in delivered_rows:
        delivered_by_month[r["delivery_date"][:7]] += r["takka_total"]

    months = sorted(set(list(given_by_month.keys()) + list(delivered_by_month.keys())))

    fabric_totals = defaultdict(float)
    for r in item_rows:
        fabric_totals[r["fabric_name"]] += r["meters_given"]

    party_totals = defaultdict(float)
    for r in item_rows:
        party_totals[r["company_name"]] += r["meters_given"]
    top_parties = sorted(party_totals.items(), key=lambda x: x[1], reverse=True)[:8]

    conn.close()

    return {
        "kpis": {
            "given": round(total_given, 2),
            "delivered": round(total_delivered, 2),
            "shortage": shortage,
            "pending": round(pending, 2)
        },
        "months": months,
        "given_by_month": [round(given_by_month.get(m, 0), 2) for m in months],
        "delivered_by_month": [round(delivered_by_month.get(m, 0), 2) for m in months],
        "fabric_labels": list(fabric_totals.keys()),
        "fabric_values": [round(v, 2) for v in fabric_totals.values()],
        "status_labels": ["Paid", "Partial", "Pending"],
        "status_values": [status_counts["paid"], status_counts["partial"], status_counts["pending"]],
        "party_labels": [p[0] for p in top_parties],
        "party_values": [round(p[1], 2) for p in top_parties]
    }

@app.route("/api/search-companies")
def search_companies():
    query = request.args.get("q", "")
    conn = get_db_connection()
    all_companies = [row["company_name"] for row in conn.execute("SELECT DISTINCT company_name FROM parties").fetchall()]
    conn.close()

    if not query:
        return {"matches": []}

    results = process.extract(query, all_companies, scorer=fuzz.WRatio, limit=5)
    matches = [r[0] for r in results if r[1] > 50]
    return {"matches": matches}


@app.route("/api/search-fabrics")
def search_fabrics():
    query = request.args.get("q", "").strip().lower()
    conn = get_db_connection()
    all_fabrics = [row["fabric_name"] for row in conn.execute("SELECT fabric_name FROM fabric_master ORDER BY fabric_name").fetchall()]
    conn.close()

    if not query:
        return {"matches": []}

    matches = []
    for fabric_name in all_fabrics:
        words = fabric_name.replace("/", " ").replace("(", " ").replace(")", " ").split()
        if any(word.lower().startswith(query) for word in words):
            matches.append(fabric_name)

    return {"matches": matches[:8]}


@app.route("/api/delivery-info/<int:delivery_id>")
def delivery_info(delivery_id):
    conn = get_db_connection()
    row = conn.execute("""
        SELECT ii.rate_per_meter,
               COALESCE(SUM(dt.meters), 0) AS takka_total
        FROM deliveries d
        JOIN inward_items ii ON d.inward_item_id = ii.id
        LEFT JOIN delivery_takkas dt ON dt.delivery_id = d.id
        WHERE d.id = ?
        GROUP BY d.id
    """, (delivery_id,)).fetchone()
    conn.close()

    if not row:
        return {"takka_total": 0, "rate_per_meter": 0, "base_amount": 0}

    rate = row["rate_per_meter"] or 0
    base_amount = row["takka_total"] * rate
    return {
        "takka_total": row["takka_total"],
        "rate_per_meter": rate,
        "base_amount": round(base_amount, 2)
    }


@app.route("/payments/<int:payment_id>/history")
def payment_history_detail(payment_id):
    conn = get_db_connection()

    payment = conn.execute("""
        SELECT pay.*, p.company_name, d.outward_challan_number
        FROM payments pay
        JOIN deliveries d ON pay.delivery_id = d.id
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        WHERE pay.id = ?
    """, (payment_id,)).fetchone()

    history_rows = conn.execute(
        "SELECT * FROM payment_history WHERE payment_id = ? ORDER BY paid_date DESC",
        (payment_id,)
    ).fetchall()

    history = []
    for h in history_rows:
        h_dict = dict(h)
        date_part = h_dict["paid_date"].split(" ")[0]
        y, m, d = date_part.split("-")
        h_dict["paid_date_display"] = f"{d}-{m}-{y}"
        history.append(h_dict)

    conn.close()
    return render_template("payment_history.html", payment=payment, history=history)

@app.route("/api/party-open-items/<int:party_id>")
def party_open_items(party_id):
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT ii.id, ii.fabric_name, ii.fabric_width_inches, ii.meters_given,
               COALESCE(SUM(d.meters_delivered), 0) AS delivered
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        LEFT JOIN deliveries d ON d.inward_item_id = ii.id
        WHERE ic.party_id = ?
        GROUP BY ii.id
        HAVING balance > 0
    """.replace("HAVING balance > 0", "HAVING (ii.meters_given - COALESCE(SUM(d.meters_delivered), 0)) > 0"),
        (party_id,)
    ).fetchall()
    conn.close()

    items = []
    for r in rows:
        balance = r["meters_given"] - r["delivered"]
        items.append({
            "id": r["id"],
            "fabric_name": r["fabric_name"],
            "fabric_width_inches": r["fabric_width_inches"] or "-",
            "balance": balance
        })
    return {"items": items}

@app.route("/deliveries/confirm/<challan_number>")
def delivery_confirm(challan_number):
    conn = get_db_connection()
    party_id = request.args.get("party_id")

    party = conn.execute("SELECT company_name, phone FROM parties WHERE id = ?", (party_id,)).fetchone()

    rows = conn.execute("""
        SELECT d.id AS delivery_id, ii.fabric_name, ii.fabric_width_inches,
               SUM(dt.meters) AS takka_total,
               COUNT(dt.meters) AS takka_count,
               GROUP_CONCAT(dt.meters, ', ') AS takka_list
        FROM deliveries d
        JOIN inward_items ii ON d.inward_item_id = ii.id
        LEFT JOIN delivery_takkas dt ON dt.delivery_id = d.id
        WHERE d.outward_challan_number = ?
        GROUP BY d.id
    """, (challan_number,)).fetchall()
    conn.close()

    items = []
    for r in rows:
        items.append({
            "fabric_name": r["fabric_name"],
            "size": f'{r["fabric_width_inches"]}"' if r["fabric_width_inches"] else "-",
            "takka_list": r["takka_list"] or "-",
            "takka_count": r["takka_count"] or 0,
            "quantity": round(r["takka_total"] or 0, 2)
        })

    date_display = datetime.now().strftime("%d-%m-%Y")

    pdf_challan.generate_challan_pdf(party["company_name"], challan_number, date_display, "-", items)

    items_summary = ", ".join(f"{i['fabric_name']} - {i['quantity']}m" for i in items)

    message_lines = [
        "Delivery Challan from Krupalu Creation",
        f"To: {party['company_name']}",
        f"Challan No: {challan_number}",
        "",
        "PDF challan attached separately, please attach the downloaded file here.",
    ]
    whatsapp_message = "%0A".join(line.replace(" ", "%20") for line in message_lines)

    return render_template(
        "delivery_confirm.html",
        challan_number=challan_number,
        party_name=party["company_name"],
        party_phone=party["phone"],
        items_summary=items_summary,
        whatsapp_message=whatsapp_message,
        pdf_download_url=f"/download-challan/{challan_number}"
    )


@app.route("/download-challan/<challan_number>")
def download_challan(challan_number):
    conn = get_db_connection()
    row = conn.execute("""
        SELECT p.company_name
        FROM deliveries d
        JOIN inward_items ii ON d.inward_item_id = ii.id
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        JOIN parties p ON ic.party_id = p.id
        WHERE d.outward_challan_number = ?
        LIMIT 1
    """, (challan_number,)).fetchone()
    conn.close()

    if not row:
        return "Challan not found", 404

    month_folder = datetime.now().strftime("%Y-%m")
    folder_path = os.path.join("challans", pdf_challan.safe_folder_name(row["company_name"]), month_folder)
    filename = pdf_challan.safe_folder_name(challan_number) + ".pdf"

    return send_from_directory(os.path.abspath(folder_path), filename, as_attachment=True)

@app.route("/parties/<int:party_id>/print-statement")
def print_statement(party_id):
    conn = get_db_connection()
    party = conn.execute("SELECT * FROM parties WHERE id = ?", (party_id,)).fetchone()

    selected_range = request.args.get("range", "90")
    date_filter_sql = ""
    if selected_range != "all":
        days = int(selected_range)
        cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        date_filter_sql = f"AND ic.date_received >= '{cutoff}'"

    order_rows = conn.execute(f"""
        SELECT ii.id AS item_id, ii.fabric_name, ii.fabric_width_inches AS size,
               ii.rate_per_meter AS rate, ic.date_received, ic.inward_challan_number,
               ii.meters_given,
               COALESCE(pr.printing_total, 0) AS printing_total,
               COALESCE(tk.takka_total, 0) AS delivered,
               dch.outward_challans, dch.last_delivery_date,
               COALESCE(pay.total_amount, 0) AS total_amount,
               COALESCE(pay.amount_paid, 0) AS amount_paid,
               pay.status AS payment_status
        FROM inward_items ii
        JOIN inward_challans ic ON ii.inward_challan_id = ic.id
        LEFT JOIN (
            SELECT inward_item_id, SUM(meters_delivered) AS printing_total
            FROM deliveries GROUP BY inward_item_id
        ) pr ON pr.inward_item_id = ii.id
        LEFT JOIN (
            SELECT d.inward_item_id, SUM(dt.meters) AS takka_total
            FROM deliveries d
            JOIN delivery_takkas dt ON dt.delivery_id = d.id
            GROUP BY d.inward_item_id
        ) tk ON tk.inward_item_id = ii.id
        LEFT JOIN (
            SELECT inward_item_id,
                   GROUP_CONCAT(DISTINCT outward_challan_number) AS outward_challans,
                   MAX(delivery_date) AS last_delivery_date
            FROM deliveries GROUP BY inward_item_id
        ) dch ON dch.inward_item_id = ii.id
        LEFT JOIN (
            SELECT d.inward_item_id, SUM(p.total_amount) AS total_amount,
                   SUM(p.amount_paid) AS amount_paid,
                   CASE WHEN SUM(p.amount_paid) >= SUM(p.total_amount) AND SUM(p.total_amount) > 0 THEN 'paid'
                        WHEN SUM(p.amount_paid) > 0 THEN 'partial'
                        ELSE 'pending' END AS status
            FROM payments p
            JOIN deliveries d ON p.delivery_id = d.id
            GROUP BY d.inward_item_id
        ) pay ON pay.inward_item_id = ii.id
        WHERE ic.party_id = ? {date_filter_sql}
        ORDER BY ic.date_received DESC
    """, (party_id,)).fetchall()

    orders = []
    item_ids_for_history = [o["item_id"] for o in order_rows]
    payment_history_map = {}
    delivery_details_map = {}
    if item_ids_for_history:
        placeholders_d = ",".join("?" * len(item_ids_for_history))
        delivery_rows = conn.execute(f"""
            SELECT d.inward_item_id, d.outward_challan_number, d.delivery_date,
                   COALESCE((SELECT SUM(dt.meters) FROM delivery_takkas dt WHERE dt.delivery_id = d.id), 0) AS meters
            FROM deliveries d
            WHERE d.inward_item_id IN ({placeholders_d})
            ORDER BY d.delivery_date
        """, item_ids_for_history).fetchall()
        for d in delivery_rows:
            key = d["inward_item_id"]
            if key not in delivery_details_map:
                delivery_details_map[key] = []
            y, m, day = d["delivery_date"].split("-")
            delivery_details_map[key].append({
                "challan": d["outward_challan_number"],
                "date": f"{day}-{m}-{y}",
                "meters": round(d["meters"], 2)
            })
    if item_ids_for_history:
        placeholders = ",".join("?" * len(item_ids_for_history))
        history_rows = conn.execute(f"""
            SELECT d.inward_item_id, ph.paid_amount, ph.paid_date
            FROM payment_history ph
            JOIN payments p ON ph.payment_id = p.id
            JOIN deliveries d ON p.delivery_id = d.id
            WHERE d.inward_item_id IN ({placeholders})
            ORDER BY ph.paid_date
        """, item_ids_for_history).fetchall()
        for h in history_rows:
            key = h["inward_item_id"]
            if key not in payment_history_map:
                payment_history_map[key] = []
            date_part = h["paid_date"].split(" ")[0]
            y, m, d = date_part.split("-")
            payment_history_map[key].append({"date": f"{d}-{m}-{y}", "amount": h["paid_amount"]})

    for o in order_rows:
        shortage = round(o["printing_total"] - o["delivered"], 2)
        remaining = round(o["total_amount"] - o["amount_paid"], 2)
        orders.append({
            "item_id": o["item_id"],
            "fabric_name": o["fabric_name"],
            "size": f'{o["size"]}"' if o["size"] else "-",
            "rate": o["rate"] if o["rate"] else "-",
            "date_received": o["date_received"],
            "inward_challan_number": o["inward_challan_number"],
            "delivery_details": delivery_details_map.get(o["item_id"], []),
            "meters_given": o["meters_given"],
            "delivered": round(o["delivered"], 2),
            "shortage": shortage,
            "total_amount": round(o["total_amount"], 2),
            "amount_paid": round(o["amount_paid"], 2),
            "remaining": remaining,
            "payment_status": o["payment_status"] or "no payment yet",
            "payment_history": payment_history_map.get(o["item_id"], [])
        })

    conn.close()

    total_given = sum(o["meters_given"] for o in orders)
    total_delivered = sum(o["delivered"] for o in orders)

    period_labels = {"30": "This month", "90": "Last 3 months", "180": "Last 6 months", "365": "Last 12 months", "all": "Since first entry"}
    period_label = period_labels.get(selected_range, selected_range)

    filepath, filename = pdf_challan.generate_statement_pdf(
        party["company_name"], period_label, orders, total_given, total_delivered
    )

    with open("config.json") as f:
        config = json.load(f)
    owner_phone = config.get("owner_phone", "")

    message_lines = [
        f"Statement for {party['company_name']}",
        f"Period: {period_label}",
        f"Total given: {total_given}m, Delivered: {total_delivered}m",
        "",
        "PDF attached separately - please attach the downloaded file.",
    ]
    whatsapp_message = "%0A".join(line.replace(" ", "%20") for line in message_lines)

    return render_template(
        "statement_confirm.html",
        party_name=party["company_name"],
        period_label=period_label,
        owner_phone=owner_phone,
        whatsapp_message=whatsapp_message,
        pdf_download_url=f"/download-statement/{party['company_name']}/{filename}"
    )


@app.route("/download-statement/<company_name>/<filename>")
def download_statement(company_name, filename):
    folder_path = os.path.join("statements", pdf_challan.safe_folder_name(company_name))
    return send_from_directory(os.path.abspath(folder_path), filename, as_attachment=True)

@app.route("/balance-sheet")
def balance_sheet():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT p.id, p.company_name, p.phone,
               COALESCE(SUM(pay.total_amount), 0) AS total_billed,
               COALESCE(SUM(pay.amount_paid), 0) AS total_paid
        FROM parties p
        LEFT JOIN inward_challans ic ON ic.party_id = p.id
        LEFT JOIN inward_items ii ON ii.inward_challan_id = ic.id
        LEFT JOIN deliveries d ON d.inward_item_id = ii.id
        LEFT JOIN payments pay ON pay.delivery_id = d.id
        GROUP BY p.id
        ORDER BY (COALESCE(SUM(pay.total_amount), 0) - COALESCE(SUM(pay.amount_paid), 0)) DESC
    """).fetchall()
    conn.close()

    balances = []
    grand_pending = 0
    for r in rows:
        balance = round(r["total_billed"] - r["total_paid"], 2)
        grand_pending += balance
        balances.append({
            "id": r["id"],
            "company_name": r["company_name"],
            "phone": r["phone"],
            "total_billed": round(r["total_billed"], 2),
            "total_paid": round(r["total_paid"], 2),
            "balance": balance
        })

    return render_template("balance_sheet.html", balances=balances, grand_pending=round(grand_pending, 2))

@app.route("/api/company-phone")
def company_phone():
    name = request.args.get("name", "").strip()
    conn = get_db_connection()
    row = conn.execute("SELECT phone FROM parties WHERE company_name = ?", (name,)).fetchone()
    conn.close()
    return {"phone": row["phone"] if row else ""}

@app.route("/clear-data", methods=["GET", "POST"])
def clear_data():
    with open("config.json") as f:
        config = json.load(f)
        config["admin_password"] = os.environ.get("ADMIN_PASSWORD", "")

    if request.method == "POST":
        entered_password = request.form.get("password", "")
        if entered_password != config.get("admin_password", ""):
            return render_template("clear_data.html", error="Galat password, dobara try karo.")

        conn = get_db_connection()
        tables_to_clear = [
            "payment_history", "payments", "delivery_takkas", "deliveries",
            "fabric_damage", "inward_items", "inward_challans", "parties",
            "audit_log", "reminders_log"
        ]
        for table in tables_to_clear:
            conn.execute(f"DELETE FROM {table}")
        conn.commit()
        conn.close()

        flash("Saara data clear ho gaya. Fresh start!")
        return redirect("/")

    return render_template("clear_data.html", error=None)

@app.route("/api/next-challan-number")
def next_challan_number():
    conn = get_db_connection()
    row = conn.execute("SELECT COUNT(*) AS cnt FROM (SELECT DISTINCT outward_challan_number FROM deliveries)").fetchone()
    conn.close()
    next_num = (row["cnt"] or 0) + 1
    return {"challan_number": f"KC-{next_num:02d}"}

@app.route("/api/recent-challans")
def recent_challans():
    conn = get_db_connection()
    rows = conn.execute("SELECT DISTINCT outward_challan_number FROM deliveries ORDER BY id DESC LIMIT 10").fetchall()
    conn.close()
    return {"challans": [r["outward_challan_number"] for r in rows]}

@app.route("/login", methods=["GET", "POST"])
def login():
    # Agar user ne form submit kiya hai
    if request.method == "POST":
        password = request.form.get("password")
        # config.json se password match karein
        with open("config.json") as f:
            import json
            config = json.load(f)
            config["login_password"] = os.environ.get("LOGIN_PASSWORD", "")
            
        if password == config.get("login_password"):
            session["logged_in"] = True
            return redirect("/")
        else:
            return render_template("login.html", error="Galat password. Kripya dubara try karein.")
            
    # Agar sirf page open kiya hai (GET request)
    return render_template("login.html")


@app.route("/logout")
def logout():
    # Session data ko clear karne ke liye
    session.clear() 
    return render_template("logout.html")

if __name__ == "__main__":
    app.run(debug=True)