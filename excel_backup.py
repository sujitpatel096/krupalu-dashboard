import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

BACKUP_FOLDER = "excel_backups"


def safe_filename(name):
    cleaned = "".join(c for c in name if c.isalnum() or c in (" ", "_", "-"))
    return cleaned.strip().replace(" ", "_") + ".xlsx"


def get_or_create_workbook(filepath, headers):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Log"
        ws.append(headers)
    return wb, ws


def prepend_row(ws, values):
    ws.insert_rows(2)
    for i, val in enumerate(values, start=1):
        ws.cell(row=2, column=i, value=val)


def append_order(company_name, phone, date_received, challan_number, fabric_name, width, garment_type, total_takka, meters_given, rate):
    os.makedirs(os.path.join(BACKUP_FOLDER, "Orders"), exist_ok=True)

    master_path = os.path.join(BACKUP_FOLDER, "Orders", "Master_Orders_Log.xlsx")
    master_headers = ["Date", "Company", "Phone", "Challan Number", "Fabric", "Width (in)", "Garment Type", "Total Takka", "Meters Given", "Rate/m"]
    wb, ws = get_or_create_workbook(master_path, master_headers)
    prepend_row(ws, [date_received, company_name, phone, challan_number, fabric_name, width or "-", garment_type or "-", total_takka or "-", meters_given, rate or "-"])
    wb.save(master_path)

    party_path = os.path.join(BACKUP_FOLDER, "Orders", safe_filename(company_name))
    party_headers = ["Date", "Phone", "Challan Number", "Fabric", "Width (in)", "Garment Type", "Total Takka", "Meters Given", "Rate/m"]
    wb2, ws2 = get_or_create_workbook(party_path, party_headers)
    prepend_row(ws2, [date_received, phone, challan_number, fabric_name, width or "-", garment_type or "-", total_takka or "-", meters_given, rate or "-"])
    wb2.save(party_path)


def append_delivery(company_name, phone, date_delivered, outward_challan, fabric_name, width, takka_total, takka_count, printing_meters, shortage):
    os.makedirs(os.path.join(BACKUP_FOLDER, "Deliveries"), exist_ok=True)

    master_path = os.path.join(BACKUP_FOLDER, "Deliveries", "Master_Delivery_Log.xlsx")
    master_headers = ["Date", "Company", "Phone", "Outward Challan", "Fabric", "Width (in)", "Takka Total (m)", "Total Takka (count)", "Printing Meters", "Shortage (m)"]
    wb, ws = get_or_create_workbook(master_path, master_headers)
    prepend_row(ws, [date_delivered, company_name, phone, outward_challan, fabric_name, width or "-", takka_total, takka_count, printing_meters, shortage])
    wb.save(master_path)

    party_path = os.path.join(BACKUP_FOLDER, "Deliveries", safe_filename(company_name))
    party_headers = ["Date", "Phone", "Outward Challan", "Fabric", "Width (in)", "Takka Total (m)", "Total Takka (count)", "Printing Meters", "Shortage (m)"]
    wb2, ws2 = get_or_create_workbook(party_path, party_headers)
    prepend_row(ws2, [date_delivered, phone, outward_challan, fabric_name, width or "-", takka_total, takka_count, printing_meters, shortage])
    wb2.save(party_path)


def append_payment(company_name, date, payment_type, amount, gst_percent, total, notes=""):
    os.makedirs(os.path.join(BACKUP_FOLDER, "Payments"), exist_ok=True)

    master_path = os.path.join(BACKUP_FOLDER, "Payments", "Master_Payments_Log.xlsx")
    master_headers = ["Date", "Company", "Type", "Amount", "GST %", "Total", "Notes"]
    wb, ws = get_or_create_workbook(master_path, master_headers)
    prepend_row(ws, [date, company_name, payment_type, amount, gst_percent, total, notes])
    wb.save(master_path)

    party_path = os.path.join(BACKUP_FOLDER, "Payments", safe_filename(company_name))
    party_headers = ["Date", "Type", "Amount", "GST %", "Total", "Notes"]
    wb2, ws2 = get_or_create_workbook(party_path, party_headers)
    prepend_row(ws2, [date, payment_type, amount, gst_percent, total, notes])
    wb2.save(party_path)